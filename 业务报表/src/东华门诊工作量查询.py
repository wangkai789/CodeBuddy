import pandas as pd
import os
import shutil
import time
import sys
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

data_dir = r'D:\文档\正骨医院\20260226职称数据\data'

# 钉钉Webhook URL
DINGTALK_WEBHOOK = "https://oapi.dingtalk.com/robot/send?access_token=d68e1626c9e5689513089f2d67a9b698067442884f342c733b82e953abb71df7"

# 从命令行参数获取医生姓名
if len(sys.argv) > 1:
    doctor_name = sys.argv[1]
else:
    doctor_name = input("请输入医生姓名: ").strip()

output_file = rf'D:\文档\正骨医院\20260226职称数据\{doctor_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
temp_file = rf'D:\文档\正骨医院\20260226职称数据\{doctor_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}_temp.xlsx'

files = [
    '2019医生组出入院统计报表.xlsx',
    '2020医生组出入院统计报表.xlsx',
    '2021医生组出入院统计报表.xlsx',
    '2022医生组出入院统计报表.xlsx',
    '2023医生组出入院统计报表.xlsx',
    '2023医生组出入院统计报表(康复).xlsx',
    '2024医生组出入院统计报表.xlsx',
    '2024医生组出入院统计报表(康复).xlsx',
    '2025医生组出入院统计报表.xlsx',
    '2025医生组出入院统计报表(康复).xlsx',
    '202601-04.15医生组出入院统计报表.xlsx',   
    '202601-04.15医生组出入院统计报表(康复).xlsx'
]

yearly_data = {}

for filename in files:
    filepath = os.path.join(data_dir, filename)
    try:
        # 提取期间（年度或年月）
        year = filename[:4]
        if filename.startswith('202601-04.15'):
            period = '202601~04.15'  # 2026年1-3月
        else:
            period = year  # 其他按年度
        
        if year.isdigit() and 2019 <= int(year) <= 2022:
            header_row = 2
        else:
            header_row = 4
        
        df = pd.read_excel(filepath, header=header_row)
        df.columns = [str(col).strip() for col in df.columns]
        
        column_rename = {
            '入院人数': '入院人次',
            '出院人数': '出院人次',
            '出院患者占床日': '出院患者占用床日数'
        }
        df.rename(columns=column_rename, inplace=True)
        
        doctor_col = None
        for col in df.columns:
            if col == '医生':
                doctor_col = col
                break
        
        if doctor_col:
            mask = df[doctor_col].astype(str).str.strip() == doctor_name.strip()
            if mask.any():
                matched = df[mask].copy()
                
                if period not in yearly_data:
                    yearly_data[period] = {
                        '期间': period,
                        '医师姓名': doctor_name,
                        '门诊人次': 0
                    }
                
                # 当门诊人次(不含零费用)有值时，优先使用该值
                outpatient_no_zero = pd.to_numeric(matched['门诊人次(不含零费用)'], errors='coerce')
                outpatient = pd.to_numeric(matched['门诊人次'], errors='coerce')
                
                # 如果门诊人次(不含零费用)有值，使用该值；否则使用门诊人次
                if outpatient_no_zero.notna().any() and outpatient_no_zero.sum() > 0:
                    yearly_data[period]['门诊人次'] += outpatient_no_zero.sum()
                else:
                    yearly_data[period]['门诊人次'] += outpatient.sum()
                
                print(f"  {filename}: 找到 {(~matched[doctor_col].isna()).sum()} 条记录")
    except Exception as e:
        print(f"Error reading {filename}: {e}")

if yearly_data:
    result_df = pd.DataFrame(list(yearly_data.values()))
    result_df = result_df.sort_values('期间')
    
    cols = ['期间', '医师姓名', '门诊人次']
    result_df = result_df[cols]
    
    total_row = pd.DataFrame([{
        '期间': '合计',
        '医师姓名': '',
        '门诊人次': result_df['门诊人次'].sum()
    }])
    result_df = pd.concat([result_df, total_row], ignore_index=True)
    
    if os.path.exists(temp_file):
        try:
            os.remove(temp_file)
        except:
            pass
    
    if os.path.exists(output_file):
        try:
            os.remove(output_file)
        except:
            time.sleep(1)
            try:
                os.remove(output_file)
            except:
                pass
    
    wb = Workbook()
    ws = wb.active
    
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    headers = list(result_df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, str(header))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row_idx, row in result_df.iterrows():
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row_idx + 2, col_idx, value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 16
    
    ws.row_dimensions[1].height = 20
    for row_idx in range(2, len(result_df) + 2):
        ws.row_dimensions[row_idx].height = 18
    
    wb.save(temp_file)
    
    time.sleep(1)
    # 多次尝试删除旧文件
    for _ in range(3):
        if os.path.exists(output_file):
            try:
                os.remove(output_file)
                break
            except:
                time.sleep(0.5)
        else:
            break
    
    shutil.move(temp_file, output_file)
    
    print(f"\n完成！共汇总 {len(yearly_data)} 个期间")
    print(f"结果已保存到: {output_file}")
    
    # === 钉钉消息已屏蔽 ===
    # 发送结果到钉钉
    # message = f"【{doctor_name}】医生数据汇总\n\n"
    # for _, row in result_df.iterrows():
    #     message += f"{row['期间']}: {row['医师姓名']} 门诊人次={row['门诊人次']}\n"
    # 
    # data = {
    #     "msgtype": "text",
    #     "text": {
    #         "content": message
    #     }
    # }
    # 
    # try:
    #     response = requests.post(DINGTALK_WEBHOOK, json=data, timeout=10)
    #     if response.json().get('errcode') == 0:
    #         print("结果已发送到钉钉")
    #     else:
    #         print(f"钉钉发送失败: {response.text}")
    # except Exception as e:
    #     print(f"钉钉发送异常: {e}")
        
else:
    print("未找到任何记录")
