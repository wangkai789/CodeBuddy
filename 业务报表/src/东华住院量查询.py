# 东华住院量查询
import pyodbc
import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# 数据库配置
DB_CONFIG = {
    'server': '192.168.101.19',
    'port': '1433',
    'user': 'dhuser',
    'password': 'dhcc',
    'database': 'dhcc'
}


def get_connection():
    """获取数据库连接"""
    conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={DB_CONFIG['server']},{DB_CONFIG['port']};DATABASE={DB_CONFIG['database']};UID={DB_CONFIG['user']};PWD={DB_CONFIG['password']}"
    return pyodbc.connect(conn_str)


def query_doctor_sp(name):c
    """正高职称医师工作量查询"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("EXEC dhcc..Search_Doctor_SP_Workload ?", (name,))
    columns = [column[0] for column in cursor.description]
    rows = cursor.fetchall()
    conn.close()
    return columns, rows


def query_doctor_ap(name):
    """副高职称医师工作量查询"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("EXEC dhcc..Search_Doctor_AP_Workload ?", (name,))
    columns = [column[0] for column in cursor.description]
    rows = cursor.fetchall()
    conn.close()
    return columns, rows


def query_nurse(name):
    """护士工作量查询"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("EXEC dhcc..Search_Nurse_Workload ?", (name,))
    columns = [column[0] for column in cursor.description]
    rows = cursor.fetchall()
    conn.close()
    return columns, rows


def query_anes(name):
    """麻醉工作量查询"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("EXEC dhcc..Search_Anes_Workload ?", (name,))
    columns = [column[0] for column in cursor.description]
    rows = cursor.fetchall()
    conn.close()
    return columns, rows


def save_to_excel(columns, rows, name, work_type, output_dir):
    """保存查询结果到Excel"""
    if not rows:
        print(f"未找到 {work_type} [{name}] 的数据")
        return None

    df = pd.DataFrame.from_records(rows, columns=columns)
    
    output_file = os.path.join(output_dir, f"{name}_{work_type}_{datetime.now().strftime('%Y%m%d')}.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = work_type

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 写入表头
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(1, col_idx, str(header))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 写入数据
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # 自动调整列宽
    for col_idx, header in enumerate(columns, 1):
        max_length = len(str(header))
        for row_idx in range(2, len(rows) + 2):
            cell_value = str(ws.cell(row_idx, col_idx).value or '')
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"].width = min(max_length + 2, 30)

    ws.row_dimensions[1].height = 20
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 18

    wb.save(output_file)
    return output_file


def main():
    print("=" * 50)
    print("东华住院量查询系统")
    print("=" * 50)
    print(f"服务器: {DB_CONFIG['server']}")
    print(f"数据库: {DB_CONFIG['database']}")
    print("=" * 50)
    
    print("\n查询类型:")
    print("1. 正高职称医师")
    print("2. 副高职称医师")
    print("3. 护士")
    print("4. 麻醉师")
    print("5. 查询全部类型")
    
    choice = input("\n请选择查询类型 (1-5): ").strip()
    
    name = input("请输入查询姓名: ").strip()
    if not name:
        print("姓名不能为空！")
        return
    
    # 输出目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(os.path.dirname(script_dir), '输出')
    os.makedirs(output_dir, exist_ok=True)
    
    results = []
    
    try:
        if choice == '1':
            print(f"\n正在查询正高职称医师 [{name}]...")
            columns, rows = query_doctor_sp(name)
            result = save_to_excel(columns, rows, name, "正高医师", output_dir)
            if result:
                results.append(("正高医师", result))
                
        elif choice == '2':
            print(f"\n正在查询副高职称医师 [{name}]...")
            columns, rows = query_doctor_ap(name)
            result = save_to_excel(columns, rows, name, "副高医师", output_dir)
            if result:
                results.append(("副高医师", result))
                
        elif choice == '3':
            print(f"\n正在查询护士 [{name}]...")
            columns, rows = query_nurse(name)
            result = save_to_excel(columns, rows, name, "护士", output_dir)
            if result:
                results.append(("护士", result))
                
        elif choice == '4':
            print(f"\n正在查询麻醉师 [{name}]...")
            columns, rows = query_anes(name)
            result = save_to_excel(columns, rows, name, "麻醉师", output_dir)
            if result:
                results.append(("麻醉师", result))
                
        elif choice == '5':
            print(f"\n正在查询全部类型 [{name}]...")
            
            try:
                print("  查询正高职称医师...")
                columns, rows = query_doctor_sp(name)
                result = save_to_excel(columns, rows, name, "正高医师", output_dir)
                if result:
                    results.append(("正高医师", result))
            except Exception as e:
                print(f"  正高医师查询失败: {e}")
            
            try:
                print("  查询副高职称医师...")
                columns, rows = query_doctor_ap(name)
                result = save_to_excel(columns, rows, name, "副高医师", output_dir)
                if result:
                    results.append(("副高医师", result))
            except Exception as e:
                print(f"  副高医师查询失败: {e}")
            
            try:
                print("  查询护士...")
                columns, rows = query_nurse(name)
                result = save_to_excel(columns, rows, name, "护士", output_dir)
                if result:
                    results.append(("护士", result))
            except Exception as e:
                print(f"  护士查询失败: {e}")
            
            try:
                print("  查询麻醉师...")
                columns, rows = query_anes(name)
                result = save_to_excel(columns, rows, name, "麻醉师", output_dir)
                if result:
                    results.append(("麻醉师", result))
            except Exception as e:
                print(f"  麻醉师查询失败: {e}")
        else:
            print("无效的选择！")
            return
            
    except Exception as e:
        print(f"\n查询出错: {e}")
        return
    
    print("\n" + "=" * 50)
    if results:
        print("查询完成！结果已保存:")
        for work_type, file_path in results:
            print(f"  {work_type}: {file_path}")
    else:
        print("未找到任何数据")
    print("=" * 50)


if __name__ == '__main__':
    main()
