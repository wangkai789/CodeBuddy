import pandas as pd
import os
import shutil
import time
import sys
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

data_dir = r'D:\文档\正骨医院\20260401病案业务报表查询\data'
file_name = '2025医疗业务报表.xlsx'

# 钉钉Webhook URL
DINGTALK_WEBHOOK = "https://oapi.dingtalk.com/robot/send?access_token=d68e1626c9e5689513089f2d67a9b698067442884f342c733b82e953abb71df7"

# 从命令行参数获取科室名称
if len(sys.argv) > 1:
    dept_name = sys.argv[1]
else:
    dept_name = input("请输入科室名称: ").strip()

output_file = rf'D:\文档\正骨医院\20260401病案业务报表查询\{dept_name}.xlsx'
temp_file = rf'D:\文档\正骨医院\20260401病案业务报表查询\{dept_name}_temp.xlsx'

filepath = os.path.join(data_dir, file_name)

# 读取Excel，header在第4行
df = pd.read_excel(filepath, header=4)

# 清理列名
df.columns = [str(col).strip() for col in df.columns]

# 查找科室列
dept_col = None
for col in df.columns:
    if col == '科室':
        dept_col = col
        break

# 从文件名提取年度
year = ''.join(filter(str.isdigit, file_name))[:4] if file_name else '2025'

results = []

if dept_col:
    # 搜索包含关键词的科室
    mask = df[dept_col].astype(str).str.contains(dept_name, na=False)
    if mask.any():
        matched = df[mask].copy()
        
        for _, row in matched.iterrows():
            mz_num = row['门急诊人数']
            mz_nozero = row['门急诊人数(不含零费用)']
            cy_num = row['出院人数']
            
            # 处理可能的NaN和类型转换
            mz_num = 0 if pd.isna(mz_num) else int(mz_num)
            mz_nozero = 0 if pd.isna(mz_nozero) else int(mz_nozero)
            cy_num = 0 if pd.isna(cy_num) else int(cy_num)
            
            results.append({
                '年度': year,
                '科室': row[dept_col],
                '门急诊人数': mz_num,
                '门急诊人数(不含零费用)': mz_nozero,
                '出院人次': cy_num
            })
        
        print(f"找到 {len(results)} 条记录")

if results:
    result_df = pd.DataFrame(results)
    
    # 添加合计行
    total_row = pd.DataFrame([{
        '年度': '',
        '科室': '合计',
        '门急诊人数': result_df['门急诊人数'].sum(),
        '门急诊人数(不含零费用)': result_df['门急诊人数(不含零费用)'].sum(),
        '出院人次': result_df['出院人次'].sum()
    }])
    result_df = pd.concat([result_df, total_row], ignore_index=True)
    
    # 保留原始数据，不进行处理
    
    # 去掉科室字段中的前缀"A"
    result_df['科室'] = result_df['科室'].str.replace('^A', '', regex=True)
    
    # 风湿病三科名称规范化
    # 先把含全角括号的版本改为半角括号（只要改括号，不需要改前缀）
    result_df['科室'] = result_df['科室'].str.replace('风湿病三科(骨质疏松科）', '风湿病三科(骨质疏松科)', regex=False)
    # 再把不包含(骨质疏松科的"风湿病三科"加上后缀
    result_df['科室'] = result_df['科室'].str.replace(r'风湿病三科(?!\(骨质疏松科\))', r'风湿病三科(骨质疏松科)', regex=True)
    
    # 按科室汇总（排除合计行）
    result_df_no_total = result_df[result_df['科室'] != '合计'].copy()
    result_df_total = result_df[result_df['科室'] == '合计'].copy()
    
    if len(result_df_no_total) > 0:
        result_df_no_total = result_df_no_total.groupby(['年度', '科室'], as_index=False).agg({
            '门急诊人数': 'sum',
            '门急诊人数(不含零费用)': 'sum',
            '出院人次': 'sum'
        })
        # 按风湿病一科、二科、三科的数字顺序排序
        result_df_no_total = result_df_no_total.reset_index(drop=True)
        
        def get_sort_order(dept_name):
            import re
            # 匹配阿拉伯数字
            match = re.search(r'(\d+)', str(dept_name))
            if match:
                return int(match.group(1))
            # 匹配中文数字
            cn_map = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '十': 10}
            for cn, num in cn_map.items():
                if cn in dept_name:
                    return num
            return 999
        
        result_df_no_total['_sort'] = result_df_no_total['科室'].apply(get_sort_order)
        result_df_no_total = result_df_no_total.sort_values('_sort', ascending=True).drop('_sort', axis=1).reset_index(drop=True)
        result_df = pd.concat([result_df_no_total, result_df_total], ignore_index=True)
    
    if os.path.exists(temp_file):
        os.remove(temp_file)
    
    wb = Workbook()
    ws = wb.active
    
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    headers = list(result_df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, str(header))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row_idx, row in result_df.iterrows():
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row_idx + 2, col_idx, value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 12
    
    ws.row_dimensions[1].height = 20
    for row_idx in range(2, len(result_df) + 2):
        ws.row_dimensions[row_idx].height = 18
    
    wb.save(temp_file)
    
    time.sleep(1)
    # 先删除目标文件
    if os.path.exists(output_file):
        try:
            os.remove(output_file)
        except PermissionError:
            # 如果删除失败，尝试用新名称保存
            output_file = rf'D:\文档\正骨医院\20260401病案业务报表查询\{dept_name}_new.xlsx'
    
    shutil.move(temp_file, output_file)
    
    print(f"\n完成！")
    print(f"结果已保存到: {output_file}")
    
    # 发送结果到钉钉
    message = f"【{dept_name}】科室数据汇总\n\n"
    message += "| 科室 | 门急诊人数 | 门急诊人数(不含零费用) | 出院人次 |\n"
    message += "|------|-----------|---------------------|----------|\n"
    for _, row in result_df.iterrows():
        message += f"| {row['科室']} | {row['门急诊人数']} | {row['门急诊人数(不含零费用)']} | {row['出院人次']} |\n"
    
    data = {
        "msgtype": "markdown",
        "markdown": {
            "title": f"{dept_name}科室数据",
            "text": message
        }
    }
    
    try:
        response = requests.post(DINGTALK_WEBHOOK, json=data, timeout=10)
        if response.json().get('errcode') == 0:
            print("结果已发送到钉钉")
        else:
            print(f"钉钉发送失败: {response.text}")
    except Exception as e:
        print(f"钉钉发送异常: {e}")
        
else:
    print("未找到任何记录")
