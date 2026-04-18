from flask import Flask, render_template, request, jsonify
import os
import sys
import json
from datetime import datetime

app = Flask(__name__)

# 获取程序所在目录或当前工作目录
if getattr(sys, 'frozen', False):
    # exe打包后
    APP_DIR = os.path.dirname(sys.executable)
    # Flask templates路径
    TEMPLATES_DIR = os.path.join(APP_DIR, 'templates')
    if os.path.exists(TEMPLATES_DIR):
        app.template_folder = TEMPLATES_DIR
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

# 配置文件路径
CONFIG_FILE = os.path.join(APP_DIR, 'config.json')
HISTORY_FILE = os.path.join(APP_DIR, 'query_history.json')

# 默认配置
DEFAULT_CONFIG = {
    'data_dir': r'D:\文档\正骨医院\20260226职称数据\data',
    'output_dir': r'D:\文档\正骨医院\20260226职称数据'
}

# 全局配置变量
DATA_DIR = DEFAULT_CONFIG['data_dir']
OUTPUT_DIR = DEFAULT_CONFIG['output_dir']

# 查询历史记录
query_history = []


def load_config():
    """加载配置文件"""
    global DATA_DIR, OUTPUT_DIR
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
                DATA_DIR = config.get('data_dir', DEFAULT_CONFIG['data_dir'])
                OUTPUT_DIR = config.get('output_dir', DEFAULT_CONFIG['output_dir'])
        except Exception as e:
            print(f"加载配置失败: {e}")
            DATA_DIR = DEFAULT_CONFIG['data_dir']
            OUTPUT_DIR = DEFAULT_CONFIG['output_dir']


def save_config(data_dir, output_dir):
    """保存配置文件"""
    global DATA_DIR, OUTPUT_DIR
    try:
        config = {
            'data_dir': data_dir,
            'output_dir': output_dir
        }
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        DATA_DIR = data_dir
        OUTPUT_DIR = output_dir
        print(f"配置已保存: 数据目录={DATA_DIR}, 输出目录={OUTPUT_DIR}")
        return True
    except Exception as e:
        print(f"保存配置失败: {e}")
        return False


def load_history():
    """从文件加载查询历史"""
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载历史记录失败: {e}")
    return []


def save_history(history):
    """保存查询历史到文件"""
    try:
        with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"保存历史记录失败: {e}")


# 启动时加载配置和历史记录
load_config()
query_history = load_history()


def run_query(doctor_name):
    """执行工作量查询"""
    import pandas as pd
    import shutil
    import time
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    files = [
        '2019医生组出入院统计报表.xlsx',
        '2020医生组出入院统计报表.xlsx',
        '2021医生组出入院统计报表.xlsx',
        '2022医生组出入院统计报表.xlsx',
        '2023医生组出入院统计报表.xlsx',
        '2023医生组出入院统计报表(康复).xlsx',
        '2024医生组出入院统计报表.xlsx',
        '2024医生组出入院统计报表(康复).xlsx',
        '2025医生组出入院统计报表.xlsx',
        '2025医生组出入院统计报表(康复).xlsx',
        '202601-04.15医生组出入院统计报表.xlsx',
        '202601-04.15医生组出入院统计报表(康复).xlsx'
    ]

    yearly_data = {}
    messages = []
    processed_files = []

    for filename in files:
        filepath = os.path.join(DATA_DIR, filename)
        try:
            year = filename[:4]
            if filename.startswith('202601-04.15'):
                period = '202601~04.15'
            else:
                period = year

            if year.isdigit() and 2019 <= int(year) <= 2022:
                header_row = 2
            else:
                header_row = 4

            df = pd.read_excel(filepath, header=header_row)
            df.columns = [str(col).strip() for col in df.columns]

            df.rename(columns={
                '入院人数': '入院人次',
                '出院人数': '出院人次',
                '出院患者占床日': '出院患者占用床日数'
            }, inplace=True)

            doctor_col = None
            for col in df.columns:
                if col == '医生':
                    doctor_col = col
                    break

            if doctor_col:
                mask = df[doctor_col].astype(str).str.strip() == doctor_name.strip()
                if mask.any():
                    matched = df[mask].copy()
                    processed_files.append(filename)

                    if period not in yearly_data:
                        yearly_data[period] = {
                            '期间': period,
                            '医师姓名': doctor_name,
                            '门诊人次': 0
                        }

                    outpatient_no_zero = pd.to_numeric(matched['门诊人次(不含零费用)'], errors='coerce')
                    outpatient = pd.to_numeric(matched['门诊人次'], errors='coerce')

                    if outpatient_no_zero.notna().any() and outpatient_no_zero.sum() > 0:
                        yearly_data[period]['门诊人次'] += outpatient_no_zero.sum()
                    else:
                        yearly_data[period]['门诊人次'] += outpatient.sum()

                    count = (~matched[doctor_col].isna()).sum()
                    messages.append(f"{filename}: 找到 {count} 条记录")
        except Exception as e:
            messages.append(f"处理 {filename} 出错: {str(e)}")

    if yearly_data:
        result_df = pd.DataFrame(list(yearly_data.values()))
        result_df = result_df.sort_values('期间')

        cols = ['期间', '医师姓名', '门诊人次']
        result_df = result_df[cols]

        total_row = pd.DataFrame([{
            '期间': '合计',
            '医师姓名': '',
            '门诊人次': result_df['门诊人次'].sum()
        }])
        result_df = pd.concat([result_df, total_row], ignore_index=True)

        # 保存到xls子目录
        xls_dir = os.path.join(OUTPUT_DIR, 'xls')
        if not os.path.exists(xls_dir):
            os.makedirs(xls_dir, exist_ok=True)
        temp_file = os.path.join(xls_dir, f'{doctor_name}_门诊工作量_temp.xlsx')
        output_file = os.path.join(xls_dir, f'{doctor_name}_门诊工作量.xlsx')

        if os.path.exists(temp_file):
            os.remove(temp_file)

        wb = Workbook()
        ws = wb.active

        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = list(result_df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, str(header))
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for row_idx, row in result_df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row_idx + 2, col_idx, value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 16

        ws.row_dimensions[1].height = 20
        for row_idx in range(2, len(result_df) + 2):
            ws.row_dimensions[row_idx].height = 18

        wb.save(temp_file)
        time.sleep(0.5)

        if os.path.exists(output_file):
            try:
                os.remove(output_file)
            except:
                pass
        shutil.move(temp_file, output_file)

        messages.append(f"完成！共汇总 {len(yearly_data)} 个期间")
        messages.append(f"结果已保存到: {output_file}")

        # 记录查询历史（更新已有记录或新增）
        history_record = {
            'doctor_name': doctor_name,
            'output_file': f'{doctor_name}_门诊工作量.xlsx',
            'output_path': output_file,
            'query_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'periods_count': len(yearly_data),
            'processed_files': processed_files
        }
        # 检查是否已存在相同医师的记录
        existing_index = None
        for i, record in enumerate(query_history):
            if record.get('doctor_name') == doctor_name:
                existing_index = i
                break
        if existing_index is not None:
            # 更新已有记录，查询次数+1
            history_record['query_count'] = query_history[existing_index].get('query_count', 1) + 1
            query_history[existing_index] = history_record
            # 移到最前面
            query_history.pop(existing_index)
            query_history.insert(0, history_record)
        else:
            # 新增记录，查询次数为1
            history_record['query_count'] = 1
            query_history.insert(0, history_record)
            # 最多保留200条记录
            if len(query_history) > 200:
                query_history.pop()
        # 保存到文件
        save_history(query_history)

        return True, output_file, messages
    else:
        return False, None, ["未找到任何记录"]


@app.route('/')
def index():
    return render_template('doctor_workload.html')


@app.route('/config')
def get_config():
    """获取当前配置"""
    return jsonify({
        'data_dir': DATA_DIR,
        'output_dir': OUTPUT_DIR
    })


@app.route('/save-config', methods=['POST'])
def update_config():
    """保存配置"""
    data_dir = request.form.get('data_dir', '').strip()
    output_dir = request.form.get('output_dir', '').strip()

    if not data_dir or not output_dir:
        return jsonify({'success': False, 'message': '数据目录和输出目录不能为空'})

    # 验证目录
    if not os.path.exists(data_dir):
        return jsonify({'success': False, 'message': '数据目录不存在'})

    # 确保输出目录存在
    try:
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
    except Exception as e:
        return jsonify({'success': False, 'message': f'创建输出目录失败: {str(e)}'})

    success = save_config(data_dir, output_dir)
    if success:
        return jsonify({'success': True, 'message': '配置保存成功'})
    else:
        return jsonify({'success': False, 'message': '配置保存失败'})


@app.route('/query', methods=['POST'])
def query():
    doctor_name = request.form.get('doctor_name', '').strip()

    if not doctor_name:
        return jsonify({'success': False, 'message': '请输入医师姓名'})

    success, output_file, messages = run_query(doctor_name)

    if success and output_file:
        # 自动生成PDF
        try:
            xlsx_to_pdf(output_file)
        except Exception as e:
            print(f"生成PDF失败: {e}")

    return jsonify({
        'success': success,
        'message': '<br>'.join(messages),
        'output_file': output_file
    })


@app.route('/delete-and-query', methods=['POST'])
def delete_and_query():
    """删除旧记录并继续查询"""
    doctor_name = request.form.get('doctor_name', '').strip()

    if not doctor_name:
        return jsonify({'success': False, 'message': '请输入医师姓名'})

    # 删除历史记录中的旧记录
    for i in range(len(query_history) - 1, -1, -1):
        if query_history[i].get('doctor_name') == doctor_name:
            record = query_history[i]
            # 删除旧文件
            try:
                if 'output_path' in record and record['output_path']:
                    if os.path.exists(record['output_path']):
                        os.remove(record['output_path'])
                    pdf_path = record['output_path'].replace('\\xls\\', '\\pdf\\').replace('/xls/', '/pdf/').replace('.xlsx', '.pdf')
                    if os.path.exists(pdf_path):
                        os.remove(pdf_path)
            except Exception as e:
                print(f"删除旧文件失败: {e}")
            query_history.pop(i)

    save_history(query_history)

    # 执行查询
    success, output_file, messages = run_query(doctor_name)

    if success and output_file:
        # 自动生成PDF
        try:
            xlsx_to_pdf(output_file)
        except Exception as e:
            print(f"生成PDF失败: {e}")

    return jsonify({
        'success': success,
        'message': '<br>'.join(messages),
        'output_file': output_file
    })


@app.route('/history')
def get_history():
    """获取查询历史"""
    return jsonify({'history': query_history})


@app.route('/open-dir')
def open_dir():
    """打开指定目录"""
    dir_path = request.args.get('path', '')
    try:
        if dir_path and os.path.exists(dir_path):
            os.startfile(dir_path)
            return jsonify({'success': True, 'path': dir_path})
        else:
            return jsonify({'success': False, 'message': f'目录不存在: {dir_path}'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/open-output-dir')
def open_output_dir():
    """打开输出目录"""
    try:
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR, exist_ok=True)
        print(f"正在打开目录: {OUTPUT_DIR}")
        os.startfile(OUTPUT_DIR)
        return jsonify({'success': True, 'path': OUTPUT_DIR})
    except Exception as e:
        print(f"打开目录失败: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/clear-history', methods=['POST'])
def clear_history():
    """清除查询历史"""
    global query_history
    # 删除对应的文件
    for record in query_history:
        if 'output_path' in record and record['output_path']:
            try:
                # 删除xlsx文件
                if os.path.exists(record['output_path']):
                    os.remove(record['output_path'])
                # 同时删除PDF文件（在pdf目录下）
                pdf_path = record['output_path'].replace('\\xls\\', '\\pdf\\').replace('/xls/', '/pdf/').replace('.xlsx', '.pdf')
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
            except Exception as e:
                print(f"删除文件失败: {e}")
    query_history = []
    save_history(query_history)
    return jsonify({'success': True})


@app.route('/delete-record', methods=['POST'])
def delete_record():
    """删除单条历史记录及对应文件"""
    global query_history
    data = request.get_json()
    index = data.get('index')
    file_path = data.get('file_path')
    
    if index is not None and 0 <= index < len(query_history):
        record = query_history[index]
        # 删除文件
        try:
            # 删除xlsx文件
            if 'output_path' in record and record['output_path'] and os.path.exists(record['output_path']):
                os.remove(record['output_path'])
            # 同时删除PDF文件（在pdf目录下）
            if file_path:
                pdf_path = file_path.replace('\\xls\\', '\\pdf\\').replace('/xls/', '/pdf/').replace('.xlsx', '.pdf')
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
        except Exception as e:
            print(f"删除文件失败: {e}")
        # 从历史记录中删除
        query_history.pop(index)
        save_history(query_history)
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': '记录不存在'})


def xlsx_to_pdf(xlsx_path):
    """将xlsx文件转换为PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import openpyxl
    
    if not os.path.exists(xlsx_path):
        return None
    
    try:
        pdfmetrics.registerFont(TTFont('SimHei', 'C:/Windows/Fonts/simhei.ttf'))
        
        # 创建自定义样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontName='SimHei', fontSize=16, alignment=1)
        
        # 读取xlsx文件
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        # 获取数据
        table_data = []
        for row in ws.iter_rows(values_only=True):
            table_data.append([str(cell) if cell is not None else '' for cell in row])
        
        if not table_data:
            return None
        
        # 计算页面方向
        col_count = len(table_data[0]) if table_data else 0
        page_size = landscape(A4) if col_count > 5 else A4
        
        # 生成PDF
        xlsx_name = os.path.basename(xlsx_path).replace('.xlsx', '')
        pdf_filename = f'{xlsx_name}.pdf'
        pdf_path = os.path.join(OUTPUT_DIR, 'pdf', pdf_filename)
        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
        
        doc = SimpleDocTemplate(pdf_path, pagesize=page_size)
        elements = []
        
        # 添加标题
        elements.append(Paragraph(xlsx_name, title_style))
        elements.append(Paragraph('<br/>', styles['Normal']))
        
        # 计算列宽
        col_widths = []
        if col_count > 0:
            avg_width = (page_size[0] - 80) / col_count
            col_widths = [avg_width] * col_count
        
        # 创建表格
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), 'SimHei'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        elements.append(table)
        
        doc.build(elements)
        return pdf_path
        
    except Exception as e:
        print(f"导出PDF错误: {e}")
        import traceback
        traceback.print_exc()
        return None


@app.route('/export-history-pdf')
def export_history_pdf():
    """导出查询历史为PDF"""
    if not query_history:
        return "没有历史记录可导出", 400
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import io
        from flask import make_response
        
        pdfmetrics.registerFont(TTFont('SimHei', 'C:/Windows/Fonts/simhei.ttf'))
        
        # 创建自定义样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontName='SimHei', fontSize=16, alignment=1)
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # 添加标题
        elements.append(Paragraph('门诊工作量查询历史记录', title_style))
        elements.append(Paragraph('<br/>', styles['Normal']))
        
        # 准备表格数据
        data = [['序号', '医师姓名', '时间', '期数']]
        for i, record in enumerate(query_history[:50], 1):  # 最多50条
            data.append([
                str(i),
                record.get('doctor_name', ''),
                record.get('query_time', ''),
                str(record.get('periods_count', 0))
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), 'SimHei'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        filename = f'query_history_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        print(f"导出PDF错误: {e}")
        import traceback
        traceback.print_exc()
        return f"导出PDF失败: {str(e)}", 500


@app.route('/export-record-pdf', methods=['POST'])
def export_record_pdf():
    """导出单条记录为PDF"""
    data = request.get_json()
    file_path = data.get('file_path')
    
    if not file_path or not os.path.exists(file_path):
        return jsonify({'success': False, 'message': '文件不存在'})
    
    try:
        pdf_path = xlsx_to_pdf(file_path)
        if pdf_path:
            return jsonify({'success': True, 'message': f'PDF已导出到: {pdf_path}', 'pdf_path': pdf_path})
        else:
            return jsonify({'success': False, 'message': 'Excel文件为空'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'导出PDF失败: {str(e)}'})


@app.route('/open-pdf')
def open_pdf():
    """打开单个PDF文件"""
    file_path = request.args.get('file', '')
    if not file_path:
        return jsonify({'success': False, 'message': '文件路径为空'})
    try:
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'PDF文件不存在'})
        os.startfile(file_path)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


if __name__ == '__main__':
    print("=" * 50)
    print("东华工作量查询 Web服务")
    print(f"程序目录: {APP_DIR}")
    print(f"数据目录: {DATA_DIR}")
    print(f"输出目录: {OUTPUT_DIR}")
    print(f"历史记录: {HISTORY_FILE}")
    print(f"已加载 {len(query_history)} 条历史记录")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5002, debug=True)
