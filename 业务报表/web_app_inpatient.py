from flask import Flask, render_template, request, jsonify, make_response
import os
import sys
import json
from datetime import datetime

app = Flask(__name__)

# 获取程序所在目录或当前工作目录
if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
    # Flask templates路径
    TEMPLATES_DIR = os.path.join(APP_DIR, 'templates')
    if os.path.exists(TEMPLATES_DIR):
        app.template_folder = TEMPLATES_DIR
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

# 配置文件路径
CONFIG_FILE = os.path.join(APP_DIR, 'config_inpatient.json')
HISTORY_FILE = os.path.join(APP_DIR, 'query_history_inpatient.json')

# 默认配置
DEFAULT_CONFIG = {
    'server': '192.168.101.19',
    'port': '1433',
    'user': 'dhuser',
    'password': 'dhcc',
    'database': 'dhcc',
    'output_dir': r'D:\文档\正骨医院\住院量查询'
}

# 全局配置变量
DB_CONFIG = {}
OUTPUT_DIR = DEFAULT_CONFIG['output_dir']

# 查询历史记录
query_history = []


def load_config():
    """加载配置文件"""
    global DB_CONFIG, OUTPUT_DIR
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
                DB_CONFIG = {
                    'server': config.get('server', DEFAULT_CONFIG['server']),
                    'port': config.get('port', DEFAULT_CONFIG['port']),
                    'user': config.get('user', DEFAULT_CONFIG['user']),
                    'password': config.get('password', DEFAULT_CONFIG['password']),
                    'database': config.get('database', DEFAULT_CONFIG['database'])
                }
                OUTPUT_DIR = config.get('output_dir', DEFAULT_CONFIG['output_dir'])
        except Exception as e:
            print(f"加载配置失败: {e}")
            init_db_config()
    else:
        init_db_config()


def init_db_config():
    """初始化数据库配置"""
    global DB_CONFIG
    DB_CONFIG = {
        'server': DEFAULT_CONFIG['server'],
        'port': DEFAULT_CONFIG['port'],
        'user': DEFAULT_CONFIG['user'],
        'password': DEFAULT_CONFIG['password'],
        'database': DEFAULT_CONFIG['database']
    }


def save_config_to_file(server, port, user, password, database, output_dir):
    """保存配置文件"""
    global DB_CONFIG, OUTPUT_DIR
    try:
        config = {
            'server': server,
            'port': port,
            'user': user,
            'password': password,
            'database': database,
            'output_dir': output_dir
        }
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        DB_CONFIG = {
            'server': server,
            'port': port,
            'user': user,
            'password': password,
            'database': database
        }
        OUTPUT_DIR = output_dir
        print(f"配置已保存")
        return True
    except Exception as e:
        print(f"保存配置失败: {e}")
        return False


def load_history():
    """从文件加载查询历史"""
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载历史记录失败: {e}")
    return []


def save_history(history):
    """保存查询历史到文件"""
    try:
        with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"保存历史记录失败: {e}")


load_config()
query_history = load_history()


def get_connection():
    """获取数据库连接"""
    import pyodbc
    conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={DB_CONFIG['server']},{DB_CONFIG['port']};DATABASE={DB_CONFIG['database']};UID={DB_CONFIG['user']};PWD={DB_CONFIG['password']}"
    return pyodbc.connect(conn_str)


def query_data(name, query_type):
    """执行查询"""
    import pyodbc
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    import traceback
    
    try:
        conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={DB_CONFIG['server']},{DB_CONFIG['port']};DATABASE={DB_CONFIG['database']};UID={DB_CONFIG['user']};PWD={DB_CONFIG['password']}"
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        
        # 根据类型选择存储过程（统一使用dhcc数据库）
        procedures = {
            'sp': "EXEC dhcc..Search_Doctor_SP_Workload ?",
            'ap': "EXEC dhcc..Search_Doctor_AP_Workload ?",
            'nurse': "EXEC dhcc..Search_Nurse_Workload ?",
            'anes': "EXEC dhcc..Search_Anes_Workload ?"
        }
        
        proc_name = {
            'sp': '正高医师',
            'ap': '副高医师',
            'nurse': '护士',
            'anes': '麻醉师'
        }
        
        print(f"执行查询: {proc_name[query_type]} - {name}")
        cursor.execute(procedures[query_type], (name,))
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        conn.close()
    except pyodbc.Error as e:
        print(f"数据库错误: {e}")
        traceback.print_exc()
        return None, None, f"数据库错误: {str(e)}"
    except Exception as e:
        print(f"查询错误: {e}")
        traceback.print_exc()
        return None, None, f"查询错误: {str(e)}"
    
    # 检查是否有有效数据
    if not rows:
        return None, None, f"未找到 {proc_name[query_type]} [{name}] 的数据"
    
    # 护士查询需要检查质控护士和责任护士列
    if query_type == 'nurse':
        # 找到质控护士和责任护士的列索引
        quality_col_idx = None
        duty_col_idx = None
        for idx, col in enumerate(columns):
            if '质控' in str(col):
                quality_col_idx = idx
            if '责任' in str(col):
                duty_col_idx = idx
        
        # 检查这两列是否全部为0
        if quality_col_idx is not None or duty_col_idx is not None:
            all_zero = True
            for row in rows:
                if quality_col_idx is not None and row[quality_col_idx] and row[quality_col_idx] != 0:
                    all_zero = False
                    break
                if duty_col_idx is not None and row[duty_col_idx] and row[duty_col_idx] != 0:
                    all_zero = False
                    break
            
            if all_zero and rows:
                return None, None, f"未找到 {proc_name[query_type]} [{name}] 的有效数据（质控护士/责任护士数据为0）"
    
    # 正高职称医师查询需要检查关键列
    if query_type == 'sp':
        key_cols = []  # 存储关键列索引
        for idx, col in enumerate(columns):
            col_str = str(col)
            if '主任' in col_str or '副主任' in col_str:
                if '出院' in col_str:
                    key_cols.append(idx)
            if '主治' in col_str:
                if '出院' in col_str:
                    key_cols.append(idx)
            if '手术' in col_str and ('三四' in col_str or '操作' in col_str):
                key_cols.append(idx)
        
        # 检查这些关键列是否全部为0
        if key_cols:
            all_zero = True
            for row in rows:
                for col_idx in key_cols:
                    if row[col_idx] and row[col_idx] != 0:
                        all_zero = False
                        break
                if not all_zero:
                    break
            
            if all_zero and rows:
                return None, None, f"未找到 {proc_name[query_type]} [{name}] 的有效数据（主任/主治医师-出院人数、手术操作数据为0）"
    
    # 副高职称医师查询需要检查关键列
    if query_type == 'ap':
        key_cols = []  # 存储关键列索引
        for idx, col in enumerate(columns):
            col_str = str(col)
            if '主治' in col_str:
                if '出院' in col_str:
                    key_cols.append(idx)
            if '住院' in col_str:
                if '出院' in col_str:
                    key_cols.append(idx)
            if '手术' in col_str and ('三四' in col_str or '操作' in col_str):
                key_cols.append(idx)
        
        # 检查这些关键列是否全部为0
        if key_cols:
            all_zero = True
            for row in rows:
                for col_idx in key_cols:
                    if row[col_idx] and row[col_idx] != 0:
                        all_zero = False
                        break
                if not all_zero:
                    break
            
            if all_zero and rows:
                return None, None, f"未找到 {proc_name[query_type]} [{name}] 的有效数据（主治/住院医师-出院人数、手术操作数据为0）"
    
    # 麻醉师查询需要检查麻醉量列
    if query_type == 'anes':
        key_cols = []  # 存储关键列索引
        for idx, col in enumerate(columns):
            col_str = str(col)
            if '麻醉' in col_str:
                key_cols.append(idx)
        
        # 检查麻醉量列是否全部为0
        if key_cols:
            all_zero = True
            for row in rows:
                for col_idx in key_cols:
                    if row[col_idx] and row[col_idx] != 0:
                        all_zero = False
                        break
                if not all_zero:
                    break
            
            if all_zero and rows:
                return None, None, f"未找到 {proc_name[query_type]} [{name}] 的有效数据（麻醉量数据为0）"
    
    # 其他类型查询：只要有数据就输出文件
    if not rows:
        return None, None, f"未找到 {proc_name[query_type]} [{name}] 的数据"
    
    # 保存到Excel（xls子目录）
    xls_dir = os.path.join(OUTPUT_DIR, 'xls')
    if not os.path.exists(xls_dir):
        os.makedirs(xls_dir, exist_ok=True)
    
    output_file = os.path.join(xls_dir, f"{name}_{proc_name[query_type]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = proc_name[query_type]

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 写入表头（自动换行）
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(1, col_idx, str(header))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # 写入数据（自动换行）
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

    # 自动调整列宽（根据内容和表头自适应）
    for col_idx in range(1, len(columns) + 1):
        # 计算表头长度
        max_length = len(str(columns[col_idx - 1]))
        # 计算数据中最长内容
        for row_idx in range(2, len(rows) + 2):
            cell_value = str(ws.cell(row_idx, col_idx).value or '')
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        # 设置列宽（最小8，最大40）
        col_letter = chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"
        ws.column_dimensions[col_letter].width = max(8, min(max_length + 2, 40))

    # 设置行高（表头行更高）
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 20

    wb.save(output_file)
    
    return columns, rows, output_file


@app.route('/')
def index():
    return render_template('inpatient_query.html')


@app.route('/config')
def get_config():
    """获取当前配置"""
    return jsonify({
        'server': DB_CONFIG.get('server', ''),
        'port': DB_CONFIG.get('port', ''),
        'user': DB_CONFIG.get('user', ''),
        'database': DB_CONFIG.get('database', ''),
        'output_dir': OUTPUT_DIR
    })


@app.route('/save-config', methods=['POST'])
def update_config():
    """保存配置"""
    server = request.form.get('server', '').strip()
    port = request.form.get('port', '').strip()
    user = request.form.get('user', '').strip()
    password = request.form.get('password', '').strip()
    database = request.form.get('database', '').strip()
    output_dir = request.form.get('output_dir', '').strip()

    if not all([server, port, user, database, output_dir]):
        return jsonify({'success': False, 'message': '请填写完整配置信息'})

    # 确保输出目录存在
    try:
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
    except Exception as e:
        return jsonify({'success': False, 'message': f'创建输出目录失败: {str(e)}'})

    success = save_config_to_file(server, port, user, password, database, output_dir)
    if success:
        return jsonify({'success': True, 'message': '配置保存成功'})
    else:
        return jsonify({'success': False, 'message': '配置保存失败'})


@app.route('/query', methods=['POST'])
def query():
    name = request.form.get('name', '').strip()
    query_type = request.form.get('query_type', '').strip()

    if not name:
        return jsonify({'success': False, 'message': '请输入查询姓名'})
    
    if not query_type:
        return jsonify({'success': False, 'message': '请选择查询类型'})

    columns, rows, result = query_data(name, query_type)
    
    if columns is None:
        return jsonify({'success': False, 'message': result})
    
    # 生成PDF
    pdf_path = None
    try:
        pdf_path = xlsx_to_pdf(result)
    except Exception as e:
        print(f"生成PDF失败: {e}")
    
    # 记录历史
    type_names = {'sp': '正高医师', 'ap': '副高医师', 'nurse': '护士', 'anes': '麻醉师'}
    history_record = {
        'name': name,
        'query_type': type_names.get(query_type, query_type),
        'output_file': os.path.basename(result),
        'output_path': result,
        'query_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'row_count': len(rows)
    }
    query_history.insert(0, history_record)
    if len(query_history) > 100:
        query_history.pop()
    save_history(query_history)

    msg = f'查询完成！共 {len(rows)} 条记录'
    if pdf_path:
        msg += f'\nPDF已导出到: {pdf_path}'
    
    return jsonify({
        'success': True,
        'message': msg,
        'output_file': result,
        'pdf_path': pdf_path
    })


@app.route('/history')
def get_history():
    """获取查询历史"""
    return jsonify({'history': query_history})


@app.route('/clear-history', methods=['POST'])
def clear_history():
    """清除查询历史"""
    global query_history
    # 删除对应的文件
    for record in query_history:
        if 'output_path' in record and record['output_path']:
            try:
                if os.path.exists(record['output_path']):
                    os.remove(record['output_path'])
                # 同时删除PDF文件
                pdf_path = record['output_path'].replace('\\xls\\', '\\pdf\\').replace('/xls/', '/pdf/')
                pdf_path = pdf_path.replace('.xlsx', '.pdf')
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
            except Exception as e:
                print(f"删除文件失败: {e}")
    query_history = []
    save_history(query_history)
    return jsonify({'success': True})


@app.route('/delete-record', methods=['POST'])
def delete_record():
    """删除单条历史记录及对应文件"""
    global query_history
    data = request.get_json()
    index = data.get('index')
    file_path = data.get('file_path')
    
    if index is not None and 0 <= index < len(query_history):
        record = query_history[index]
        # 删除文件
        try:
            if 'output_path' in record and record['output_path'] and os.path.exists(record['output_path']):
                os.remove(record['output_path'])
            # 同时删除PDF文件
            if file_path:
                pdf_path = file_path.replace('\\xls\\', '\\pdf\\').replace('/xls/', '/pdf/').replace('.xlsx', '.pdf')
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
        except Exception as e:
            print(f"删除文件失败: {e}")
        # 从历史记录中删除
        query_history.pop(index)
        save_history(query_history)
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': '记录不存在'})


@app.route('/export-history-pdf')
def export_history_pdf():
    """导出查询历史为PDF"""
    import io
    
    if not query_history:
        return "没有历史记录可导出", 400
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        pdfmetrics.registerFont(TTFont('SimHei', 'C:/Windows/Fonts/simhei.ttf'))
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph('Query History Record', styles['Heading1']))
        elements.append(Paragraph(f'Export Time: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Normal']))
        elements.append(Paragraph('<br/>', styles['Normal']))
        
        data = [['No', 'Name', 'Type', 'Records', 'Query Time']]
        for i, record in enumerate(query_history, 1):
            data.append([
                str(i),
                record.get('name', ''),
                record.get('query_type', ''),
                str(record.get('row_count', '')),
                record.get('query_time', '')
            ])
        
        table = Table(data, colWidths=[40, 80, 80, 60, 130])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), 'SimHei'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        filename = f'query_history_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        print(f"导出PDF错误: {e}")
        import traceback
        traceback.print_exc()
        return f"导出PDF失败: {str(e)}", 500



def xlsx_to_pdf(xlsx_path):
    """将xlsx文件转换为PDF，返回PDF路径"""
    import re
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import openpyxl
    
    pdfmetrics.registerFont(TTFont('SimHei', 'C:/Windows/Fonts/simhei.ttf'))
    
    # 创建自定义样式（支持中文）
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName='SimHei',
        fontSize=16,
        alignment=1  # 居中
    )
    
    # 创建PDF输出目录
    pdf_dir = os.path.join(OUTPUT_DIR, 'pdf')
    if not os.path.exists(pdf_dir):
        os.makedirs(pdf_dir, exist_ok=True)
    
    # 读取xlsx文件
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    
    # 获取数据
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([str(cell) if cell is not None else '' for cell in row])
    
    if not data:
        return None
    
    # 计算页面方向和尺寸
    col_count = len(data[0]) if data else 0
    page_size = landscape(A4) if col_count > 5 else A4
    
    # 生成PDF文件路径（与xlsx同名）
    xlsx_name = os.path.basename(xlsx_path).replace('.xlsx', '')
    # 提取标题（姓名_查询类型），只保留汉字
    parts = xlsx_name.rsplit('_', 2)
    title_name = ''.join(re.findall(r'[\u4e00-\u9fa5]+', '_'.join(parts[:2]))) if len(parts) >= 3 else xlsx_name
    pdf_filename = f'{xlsx_name}.pdf'
    pdf_path = os.path.join(pdf_dir, pdf_filename)
    
    doc = SimpleDocTemplate(pdf_path, pagesize=page_size)
    elements = []
    
    # 添加标题（使用中文样式）
    elements.append(Paragraph(title_name, title_style))
    elements.append(Paragraph('<br/>', getSampleStyleSheet()['Normal']))
    
    # 计算列宽
    col_widths = []
    if col_count > 0:
        avg_width = (page_size[0] - 80) / col_count
        col_widths = [avg_width] * col_count
    
    # 创建表格
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'SimHei'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    elements.append(table)
    
    doc.build(elements)
    return pdf_path


@app.route('/export-xlsx-pdf')
def export_xlsx_pdf():
    """将xlsx文件转换为PDF"""
    file_path = request.args.get('file', '')
    
    if not file_path or not os.path.exists(file_path):
        return jsonify({'success': False, 'message': '文件不存在'})
    
    try:
        pdf_path = xlsx_to_pdf(file_path)
        if pdf_path:
            return jsonify({'success': True, 'message': f'PDF已导出到: {pdf_path}', 'pdf_path': pdf_path})
        else:
            return jsonify({'success': False, 'message': 'Excel文件为空'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'转换PDF失败: {str(e)}'})


@app.route('/download-xlsx')
def download_xlsx():
    """下载xlsx文件"""
    file_path = request.args.get('file', '')
    
    if not file_path or not os.path.exists(file_path):
        return "文件不存在", 404
    
    return send_file(file_path, as_attachment=True, download_name=os.path.basename(file_path))


@app.route('/open-output-dir')
def open_output_dir():
    """打开输出目录"""
    try:
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR, exist_ok=True)
        os.startfile(OUTPUT_DIR)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/open-pdf-dir')
def open_pdf_dir():
    """打开PDF输出目录"""
    try:
        pdf_dir = os.path.join(OUTPUT_DIR, 'pdf')
        if not os.path.exists(pdf_dir):
            os.makedirs(pdf_dir, exist_ok=True)
        os.startfile(pdf_dir)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/open-pdf')
def open_pdf():
    """打开单个PDF文件"""
    file_path = request.args.get('file', '')
    if not file_path:
        return jsonify({'success': False, 'message': '文件路径为空'})
    try:
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'PDF文件不存在'})
        os.startfile(file_path)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


if __name__ == '__main__':
    print("=" * 50)
    print("东华住院量查询 Web服务")
    print(f"服务器: {DB_CONFIG.get('server', 'N/A')}")
    print(f"数据库: {DB_CONFIG.get('database', 'N/A')}")
    print(f"输出目录: {OUTPUT_DIR}")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5003, debug=True)
